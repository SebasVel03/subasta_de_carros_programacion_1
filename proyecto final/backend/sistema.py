"""
Lógica de negocio central de la app de subastas.

Esta es la versión unificada y corregida de admin.py:
- Arregla el bug de self.comi_porcentaje (ver historial de la conversación).
- Ya no pierde fecha_inicio / fecha_fin al cargar carros.json.
- Agrega registro/login de usuarios, registrar_puja() con validaciones reales,
  cierre automático de subastas vencidas, persistencia a disco, y un set de
  métodos "obtener_..." pensados para alimentar directamente las vistas de Flet.
- Agrega un reporte financiero por usuario (ingresos como vendedor, egresos
  como comprador) y su exportación a un archivo Excel con gráfico, además de
  la confirmación de entrega física de un vehículo ya vendido.
- Agrega una bandeja de mensajes GLOBAL (todas las conversaciones del
  usuario en toda la plataforma, sin importar el carro) y el conteo total
  de mensajes sin leer, para alimentar el ícono de notificaciones de la
  barra superior (ver views/bandeja_mensajes_dialog.py).
- Agrega notificaciones del sistema (por ahora, "te superaron una puja"),
  favoritos con métodos propios (antes solo vivían como método de Usuario,
  sin wrapper con validación/tupla (ok, resultado) en esta capa), y
  calificaciones de vendedor post-venta que alimentan Usuario.reputacion.
"""

import json
import re
from datetime import datetime, timedelta, timezone

from .modelos import Usuario, Carro, Puja, Mensaje, Notificacion, Calificacion, hash_password

# El rol admin/experto ya NO se elige a mano ni depende de un código fijo en
# el código fuente (eso era inseguro y fácil de compartir sin querer). Ahora
# se detecta automáticamente a partir del FORMATO del correo con el que
# alguien se registra: nombre.admin.XX@dominio (XX = cualquier número de
# 2 dígitos; no identifica un cupo específico, solo forma parte del patrón
# exigido — ver _es_formato_admin más abajo).
#
# MAX_ADMINISTRADORES es el único freno real a cuántas cuentas admin pueden
# existir en la plataforma. Arranca bajo a propósito; el equipo puede subir
# este número cuando decida abrir más cupos — no hace falta tocar nada más
# (ni la vista de login, ni el resto de sistema.py).
MAX_ADMINISTRADORES = 5

PATRON_EMAIL_ADMIN = re.compile(r".+\.admin\.\d{2}@.+", re.IGNORECASE)


def _es_formato_admin(email: str) -> bool:
    """True si el correo sigue el patrón nombre.admin.XX@dominio (XX = 2 dígitos)."""
    return bool(PATRON_EMAIL_ADMIN.fullmatch((email or "").strip()))


class AdministradorCompraVenta:
    def __init__(self, comision_plataforma_porcentaje=0.05):
        self.usuarios = {}
        self.carros = {}
        self.pujas = []
        self.mensajes = []
        self.notificaciones = []
        self.calificaciones = []
        self.comision_porcentaje = comision_plataforma_porcentaje

    # =====================================================================
    # CARGA / PERSISTENCIA
    # =====================================================================
    def cargar_datos_desde_json(self, json_u, json_c, json_p, json_m=None, json_not=None, json_cal=None):
        """Carga el sistema a partir de strings JSON ya leídos.
        json_m (mensajes del chat), json_not (notificaciones) y json_cal
        (calificaciones) son opcionales: si no se pasan, simplemente no hay
        datos cargados de esa categoría (compatibilidad con código existente
        que solo pasaba usuarios/carros/pujas)."""
        try:
            # 1. Usuarios
            for id_u, d in json.loads(json_u).items():
                password_hash = d.get("password_hash") or hash_password(d.get("contraseña", ""))
                self.usuarios[id_u] = Usuario(
                    id_usuario=id_u,
                    nombre=d["nombre"],
                    email=d["email"],
                    password_hash=password_hash,
                    rol=d["rol"],
                    telefono=d.get("telefono", ""),
                    foto_perfil=d.get("foto_perfil"),
                    verificado=d.get("verificado", False),
                    autos_en_posesion=list(d.get("autos_en_posesion", [])),
                    reputacion=d.get("reputacion_estrellas", 0.0),
                    # BUGFIX: estos dos campos se guardaban en memoria (via
                    # Usuario.registrar_oferta / agregar_favorito / quitar_favorito)
                    # pero nunca se leían de vuelta al cargar desde disco, así que
                    # se "olvidaban" cada vez que se reiniciaba la app aunque las
                    # pujas siguieran existiendo en subastas.json. Por eso
                    # 'SUBASTAS ACTIVAS' aparecía vacía tras reiniciar el main
                    # incluso habiendo pujado antes. Ver también
                    # guardar_datos_a_archivos() más abajo, que ahora sí los escribe.
                    historial_ofertas=list(d.get("historial_ofertas", [])),
                    favoritos=list(d.get("favoritos", [])),
                    fecha_registro=d.get("fecha_registro"),
                )
            # 2. Carros (ahora SÍ guardamos fecha_inicio / fecha_fin, y los
            #    nuevos campos de imagen + verificación + confirmación de entrega)
            for c in json.loads(json_c):
                id_c = c["coche_id"]
                self.carros[id_c] = Carro(
                    id_carro=id_c,
                    vendedor_id=c["vendedor_id"],
                    marca=c["marca"],
                    modelo=c["modelo"],
                    anio=c["año"],
                    kilometraje=c["kilometraje"],
                    precio_base=c["precio_base"],
                    precio_reserva=c["precio_reserva"],
                    # Los carros de ejemplo originales no tenían estado de revisión;
                    # si no viene el campo, se asume 'activa' para no romper esos datos.
                    estado_subasta=c.get("estado_subasta", "activa"),
                    fecha_inicio=c.get("fecha_inicio"),
                    fecha_fin=c.get("fecha_fin"),
                    especificaciones=c.get("especificaciones", {}),
                    extras=c.get("caracteristicas_extra", []),
                    precio_final_venta=c.get("precio_final_venta", 0.0),
                    comprador_id=c.get("comprador_id"),
                    imagen=c.get("imagen"),
                    imagenes=c.get("imagenes"),
                    condicion_general=c.get("condicion_general"),
                    descripcion_danos=c.get("descripcion_danos", ""),
                    documentos_en_regla=c.get("documentos_en_regla", False),
                    duracion_dias=c.get("duracion_dias", 7),
                    fecha_publicacion=c.get("fecha_publicacion"),
                    motivo_rechazo=c.get("motivo_rechazo"),
                    entrega_confirmada=c.get("entrega_confirmada", False),
                    fecha_entrega_confirmada=c.get("fecha_entrega_confirmada"),
                )
            # 3. Pujas
            for id_carro, lista_pujas in json.loads(json_p).items():
                for p in lista_pujas:
                    self.pujas.append(Puja(
                        p["puja_id"], id_carro, p["usuario_id"], p["monto"],
                        p["fecha_hora"], p["metodo_pago_verificado"],
                    ))
            # 4. Mensajes del chat comprador-vendedor (opcional)
            if json_m:
                for m in json.loads(json_m):
                    self.mensajes.append(Mensaje(
                        m["id"], m["id_carro"], m["id_remitente"], m["id_destinatario"],
                        m["texto"], m.get("fecha_hora"), m.get("leido", False),
                    ))
            # 5. Notificaciones del sistema (opcional)
            if json_not:
                for n in json.loads(json_not):
                    self.notificaciones.append(Notificacion(
                        n["id"], n["id_usuario_destino"], n["tipo"], n["texto"],
                        id_carro=n.get("id_carro"), fecha_hora=n.get("fecha_hora"),
                        leido=n.get("leido", False), avisada_en_app=n.get("avisada_en_app", False),
                    ))
            # 6. Calificaciones de vendedor (opcional)
            if json_cal:
                for c in json.loads(json_cal):
                    self.calificaciones.append(Calificacion(
                        c["id"], c["id_carro"], c["id_calificador"], c["id_calificado"],
                        c["estrellas"], comentario=c.get("comentario", ""), fecha_hora=c.get("fecha_hora"),
                    ))
            return True
        except Exception as e:
            print(f"❌ Error al cargar datos: {e}")
            return False

    def cargar_datos_desde_archivos(self, ruta_usuarios, ruta_carros, ruta_pujas, ruta_mensajes=None,
                                     ruta_notificaciones=None, ruta_calificaciones=None):
        """Lee los .json desde disco (UTF-8, por los acentos/ñ) y carga el sistema.
        ruta_mensajes, ruta_notificaciones y ruta_calificaciones son opcionales
        y, si el archivo todavía no existe (proyectos más viejos que no lo
        tenían), simplemente arrancan vacíos en vez de fallar."""
        with open(ruta_usuarios, encoding="utf-8") as f:
            txt_u = f.read()
        with open(ruta_carros, encoding="utf-8") as f:
            txt_c = f.read()
        with open(ruta_pujas, encoding="utf-8") as f:
            txt_p = f.read()

        def _leer_opcional(ruta):
            if not ruta:
                return None
            try:
                with open(ruta, encoding="utf-8") as f:
                    return f.read()
            except FileNotFoundError:
                return "[]"

        txt_m = _leer_opcional(ruta_mensajes)
        txt_not = _leer_opcional(ruta_notificaciones)
        txt_cal = _leer_opcional(ruta_calificaciones)

        return self.cargar_datos_desde_json(txt_u, txt_c, txt_p, txt_m, txt_not, txt_cal)

    def guardar_datos_a_archivos(self, ruta_usuarios, ruta_carros, ruta_pujas, ruta_mensajes=None,
                                  ruta_notificaciones=None, ruta_calificaciones=None):
        """
        Persiste el estado actual de vuelta a disco.

        IMPORTANTE: a partir de aquí los usuarios se guardan con "password_hash"
        en vez de "contraseña" en texto plano. cargar_datos_desde_json() ya
        soporta ambos formatos (ver arriba), así que la migración es transparente.
        """
        usuarios_json = {
            u.id: {
                "nombre": u.nombre,
                "email": u.email,
                "password_hash": u.password_hash,
                "telefono": u.telefono,
                "foto_perfil": u.foto_perfil,
                "rol": u.rol,
                "fecha_registro": u.fecha_registro,
                "reputacion_estrellas": u.reputacion,
                "verificado": u.verificado,
                "autos_en_posesion": u.autos_en_posesion,
                # BUGFIX: antes no se guardaban estos dos campos, así que las
                # pujas y favoritos de cada usuario se "perdían" al reiniciar
                # la app (ver el mismo comentario en cargar_datos_desde_json).
                "historial_ofertas": u.historial_ofertas,
                "favoritos": u.favoritos,
            }
            for u in self.usuarios.values()
        }
        carros_json = [
            {
                "coche_id": c.id,
                "vendedor_id": c.vendedor_id,
                "marca": c.marca,
                "modelo": c.modelo,
                "año": c.anio,
                "kilometraje": c.kilometraje,
                "precio_base": c.precio_base,
                "precio_reserva": c.precio_reserva,
                "estado_subasta": c.estado_subasta,
                "fecha_inicio": c.fecha_inicio,
                "fecha_fin": c.fecha_fin,
                "especificaciones": c.especificaciones,
                "caracteristicas_extra": c.extras,
                "precio_final_venta": c.precio_final_venta,
                "comprador_id": c.comprador_id,
                "imagen": c.imagen,
                "imagenes": c.imagenes,
                "condicion_general": c.condicion_general,
                "descripcion_danos": c.descripcion_danos,
                "documentos_en_regla": c.documentos_en_regla,
                "duracion_dias": c.duracion_dias,
                "fecha_publicacion": c.fecha_publicacion,
                "motivo_rechazo": c.motivo_rechazo,
                "entrega_confirmada": c.entrega_confirmada,
                "fecha_entrega_confirmada": c.fecha_entrega_confirmada,
            }
            for c in self.carros.values()
        ]
        pujas_json = {}
        for p in self.pujas:
            pujas_json.setdefault(p.id_carro, []).append({
                "puja_id": p.id,
                "usuario_id": p.id_usuario,
                "monto": p.monto,
                "fecha_hora": p.fecha_hora,
                "metodo_pago_verificado": p.pago_verificado,
            })

        with open(ruta_usuarios, "w", encoding="utf-8") as f:
            json.dump(usuarios_json, f, ensure_ascii=False, indent=2)
        with open(ruta_carros, "w", encoding="utf-8") as f:
            json.dump(carros_json, f, ensure_ascii=False, indent=2)
        with open(ruta_pujas, "w", encoding="utf-8") as f:
            json.dump(pujas_json, f, ensure_ascii=False, indent=2)

        if ruta_mensajes:
            mensajes_json = [m.to_dict() for m in self.mensajes]
            with open(ruta_mensajes, "w", encoding="utf-8") as f:
                json.dump(mensajes_json, f, ensure_ascii=False, indent=2)

        if ruta_notificaciones:
            notificaciones_json = [n.to_dict() for n in self.notificaciones]
            with open(ruta_notificaciones, "w", encoding="utf-8") as f:
                json.dump(notificaciones_json, f, ensure_ascii=False, indent=2)

        if ruta_calificaciones:
            calificaciones_json = [c.to_dict() for c in self.calificaciones]
            with open(ruta_calificaciones, "w", encoding="utf-8") as f:
                json.dump(calificaciones_json, f, ensure_ascii=False, indent=2)

    # =====================================================================
    # AUTENTICACIÓN
    # =====================================================================
    def registrar_usuario(self, nombre, email, password, telefono=""):
        """
        El rol ya no lo elige la persona desde un selector en el formulario:
        se decide solo, del lado del backend, según el formato del correo.

          - Correo normal                              -> rol 'usuario'
            (puede comprar Y vender, sin tener que elegir una sola cosa).
          - Correo con formato nombre.admin.XX@dominio  -> rol 'admin'
            (experto que revisa subastas), siempre que todavía haya cupo
            (ver MAX_ADMINISTRADORES / _es_formato_admin arriba).
        """
        if any(u.email.lower() == email.lower() for u in self.usuarios.values()):
            return False, "Ya existe una cuenta con ese correo."

        if _es_formato_admin(email):
            admins_actuales = len([u for u in self.usuarios.values() if u.rol == "admin"])
            if admins_actuales >= MAX_ADMINISTRADORES:
                return False, (
                    f"Se alcanzó el número máximo de administradores permitidos "
                    f"({MAX_ADMINISTRADORES})."
                )
            rol = "admin"
        else:
            rol = "usuario"

        nuevo_id = f"usr_{len(self.usuarios) + 1:03d}"
        nuevo_usuario = Usuario(
            id_usuario=nuevo_id,
            nombre=nombre,
            email=email,
            password_hash=hash_password(password),
            rol=rol,
            telefono=telefono,
            # TODO: reemplazar por un flujo real de verificación (confirmación
            # de correo, revisión manual de documentos, etc.). Por ahora se
            # autoverifica al registrarse para que el equipo pueda probar
            # pujar/publicar de inmediato sin quedar bloqueado.
            verificado=True,
        )
        self.usuarios[nuevo_id] = nuevo_usuario
        return True, nuevo_usuario

    def autenticar_usuario(self, email, password):
        for usuario in self.usuarios.values():
            if usuario.email.lower() == email.lower():
                if usuario.verificar_password(password):
                    return True, usuario
                return False, "Contraseña incorrecta."
        return False, "No existe una cuenta con ese correo."

    def actualizar_perfil(self, id_usuario, nombre=None, telefono=None):
        usuario = self.usuarios.get(id_usuario)
        if not usuario:
            return False, "El usuario no existe."
        if nombre:
            usuario.nombre = nombre
        if telefono is not None:
            usuario.telefono = telefono
        return True, usuario

    def actualizar_foto_perfil(self, id_usuario, foto_perfil):
        """
        Guarda (o quita, pasando foto_perfil=None) la foto de perfil de un
        usuario. Es un método aparte de actualizar_perfil() a propósito: ahí
        un valor None en nombre/telefono significa "no toques este campo",
        mientras que acá None es un valor legítimo y buscado — "el usuario
        quitó su foto" —, así que no puede compartir esa misma convención.

        foto_perfil, igual que Carro.imagen, puede ser una URL o un string
        base64 (sin el prefijo 'data:'); la vista es responsable de validar
        tamaño antes de llegar hasta acá (ver TAMANO_MAXIMO_IMAGEN_BYTES en
        views/shared.py).
        """
        usuario = self.usuarios.get(id_usuario)
        if not usuario:
            return False, "El usuario no existe."
        usuario.foto_perfil = foto_perfil
        return True, usuario

    def cambiar_password(self, id_usuario, password_actual, password_nueva):
        usuario = self.usuarios.get(id_usuario)
        if not usuario:
            return False, "El usuario no existe."
        if not usuario.verificar_password(password_actual):
            return False, "La contraseña actual no es correcta."
        if len(password_nueva) < 6:
            return False, "La contraseña nueva debe tener al menos 6 caracteres."
        usuario.password_hash = hash_password(password_nueva)
        return True, usuario

    # =====================================================================
    # MÓDULO COMPRA (publicar un carro)
    # =====================================================================
    def recibir_carro_compra(self, id_vendedor, coche_id, marca, modelo, anio, kilometraje,
                              precio_base, precio_reserva, especificaciones, extras,
                              imagen=None, imagenes=None, condicion_general=None, descripcion_danos="",
                              documentos_en_regla=False, duracion_dias=7):
        """
        Publica un carro nuevo. A diferencia de antes, NO queda 'activa' de
        inmediato: nace en 'pendiente_revision' y un admin/experto tiene que
        aprobarla (ver aprobar_subasta) antes de que el público pueda verla
        en 'Explorar Subastas' o pujar por ella.

        Cualquier usuario verificado puede publicar — comprar y vender ya
        no son roles exclusivos entre sí (ver registrar_usuario). Incluso un
        admin puede publicar/pujar si quiere participar como usuario normal.

        imagenes: lista opcional con TODAS las fotos del vehículo (URLs y/o
        base64). Si se pasa, 'imagen' (la portada/miniatura de siempre) se
        toma de ahí; si no se pasa, se comporta como antes (una sola foto).
        Ver Carro.__init__ en modelos.py para el detalle del fallback.
        """
        vendedor = self.usuarios.get(id_vendedor)
        if not vendedor:
            return False, f"El usuario {id_vendedor} no existe."
        if not vendedor.verificado:
            return False, f"{vendedor.nombre} no está verificado. No puede publicar autos."
        if coche_id in self.carros:
            return False, f"Ya existe un carro publicado con el id {coche_id}."

        imagenes = imagenes or ([imagen] if imagen else [])
        imagen_portada = imagen or (imagenes[0] if imagenes else None)

        nuevo_carro = Carro(
            coche_id, id_vendedor, marca, modelo, anio, kilometraje, precio_base,
            precio_reserva, estado_subasta="pendiente_revision",
            especificaciones=especificaciones, extras=extras,
            imagen=imagen_portada, imagenes=imagenes, condicion_general=condicion_general,
            descripcion_danos=descripcion_danos, documentos_en_regla=documentos_en_regla,
            duracion_dias=duracion_dias,
        )
        self.carros[coche_id] = nuevo_carro
        vendedor.autos_en_posesion.append(coche_id)
        return True, nuevo_carro

    # =====================================================================
    # MÓDULO REVISIÓN (admin/experto aprueba o rechaza antes de salir al público)
    # =====================================================================
    def aprobar_subasta(self, id_carro):
        carro = self.carros.get(id_carro)
        if not carro:
            return False, "El carro no existe."
        if carro.estado_subasta != "pendiente_revision":
            return False, f"Este carro no está pendiente de revisión (estado actual: {carro.estado_subasta})."

        ahora = datetime.now(timezone.utc)
        carro.fecha_inicio = ahora.isoformat()
        carro.fecha_fin = (ahora + timedelta(days=carro.duracion_dias)).isoformat()
        carro.estado_subasta = "activa"
        carro.motivo_rechazo = None
        return True, carro

    def rechazar_subasta(self, id_carro, motivo=""):
        carro = self.carros.get(id_carro)
        if not carro:
            return False, "El carro no existe."
        if carro.estado_subasta != "pendiente_revision":
            return False, f"Este carro no está pendiente de revisión (estado actual: {carro.estado_subasta})."

        carro.estado_subasta = "rechazada"
        carro.motivo_rechazo = motivo or "No cumple los requisitos de la plataforma."
        return True, carro

    def obtener_subastas_pendientes_revision(self, excluir_vendedor_id=None):
        """
        Cola de revisión para el admin/experto, con todos los datos que necesita
        para validar el vehículo (especificaciones, fotos, estado, vendedor).

        excluir_vendedor_id: si el admin también publica carros propios (ahora
        que cualquiera puede comprar y vender), no debería poder aprobar/rechazar
        su propia subasta — eso se filtra pasando su propio id aquí.
        """
        pendientes = [c for c in self.carros.values()
                      if c.estado_subasta == "pendiente_revision" and c.vendedor_id != excluir_vendedor_id]
        pendientes.sort(key=lambda c: c.fecha_publicacion or "")

        resultado = []
        for c in pendientes:
            vendedor = self.usuarios.get(c.vendedor_id)
            d = c.to_dict()
            d["vendedor_nombre"] = vendedor.nombre if vendedor else "Desconocido"
            d["vendedor_reputacion"] = vendedor.reputacion if vendedor else 0
            d["vendedor_verificado"] = vendedor.verificado if vendedor else False
            resultado.append(d)
        return resultado

    # =====================================================================
    # MÓDULO PUJAS (antes no existía: las pujas se insertaban a mano)
    # =====================================================================
    def registrar_puja(self, id_usuario, id_carro, monto, metodo_pago_verificado=True):
        usuario = self.usuarios.get(id_usuario)
        carro = self.carros.get(id_carro)

        if not usuario:
            return False, "El usuario no existe."
        if not carro:
            return False, "El carro no existe."
        if carro.estado_subasta != "activa" or carro.esta_vencida():
            return False, f"La subasta de {carro.marca} {carro.modelo} ya no está activa."
        if not usuario.verificado:
            return False, "Debes verificar tu cuenta antes de poder pujar."
        if usuario.id == carro.vendedor_id:
            return False, "No puedes pujar por tu propio vehículo."

        pujas_validas = [p for p in self.pujas if p.id_carro == id_carro and p.pago_verificado]
        monto_minimo = max([carro.precio_base] + [p.monto for p in pujas_validas])
        if monto <= monto_minimo:
            return False, f"La puja debe ser mayor a ${monto_minimo:,}."

        nueva_puja = Puja(
            id_puja=f"puj_{id_carro}_{len(self.pujas) + 1:03d}",
            id_carro=id_carro,
            id_usuario=id_usuario,
            monto=monto,
            fecha_hora=datetime.now(timezone.utc).isoformat(),
            pago_verificado=metodo_pago_verificado,
        )
        self.pujas.append(nueva_puja)
        usuario.registrar_oferta(id_carro=id_carro, monto=monto)

        # A quién hay que avisarle "te superaron": la persona que tenía la
        # oferta en estado 'Activa' para este carro justo ANTES de esta puja
        # (a lo sumo una, porque solo puede haber un "ganador" vigente por
        # carro a la vez). Se calcula ANTES del loop de marcado de abajo,
        # que es el que la pasa a 'Superada' -- si se calculara después ya
        # no quedaría ninguna en 'Activa' para detectar. No hace falta
        # revisar cada Puja vieja de pujas_validas: alguien que ya estaba
        # 'Superada' de una ronda anterior no vuelve a necesitar aviso
        # ahora, aunque su Puja siga en la lista (es solo historial).
        id_usuario_superado = None
        for p in pujas_validas:
            if p.id_usuario == id_usuario:
                continue
            otro = self.usuarios.get(p.id_usuario)
            if not otro:
                continue
            if any(o["id_carro"] == id_carro and o["estado"] == "Activa" for o in otro.historial_ofertas):
                id_usuario_superado = p.id_usuario
                break

        # Las pujas anteriores de este mismo carro quedan "Superada"
        for p in pujas_validas:
            otro = self.usuarios.get(p.id_usuario)
            if not otro:
                continue
            for oferta in otro.historial_ofertas:
                if oferta["id_carro"] == id_carro and oferta["estado"] == "Activa":
                    oferta["estado"] = "Superada"

        if id_usuario_superado:
            self.crear_notificacion(
                id_usuario_destino=id_usuario_superado,
                tipo="puja_superada",
                texto=f"Te superaron la puja en {carro.marca} {carro.modelo} ({carro.anio}): "
                      f"la nueva puja más alta es ${monto:,.0f}.",
                id_carro=id_carro,
            )

        return True, nueva_puja

    # =====================================================================
    # MÓDULO NOTIFICACIONES (avisos del sistema, hoy solo "puja_superada")
    # =====================================================================
    def crear_notificacion(self, id_usuario_destino, tipo, texto, id_carro=None):
        """Punto único de creación de notificaciones -- lo usa registrar_puja()
        arriba, y a futuro cualquier otro evento que necesite avisarle algo a
        un usuario (subasta aprobada/rechazada, etc.) puede llamarlo igual."""
        nueva = Notificacion(
            id_notificacion=f"not_{len(self.notificaciones) + 1:05d}",
            id_usuario_destino=id_usuario_destino,
            tipo=tipo,
            texto=texto,
            id_carro=id_carro,
        )
        self.notificaciones.append(nueva)
        return nueva

    def obtener_notificaciones_usuario(self, id_usuario):
        """Historial completo de notificaciones de este usuario, más
        recientes primero -- para el panel de views/notificaciones_dialog.py."""
        propias = [n for n in self.notificaciones if n.id_usuario_destino == id_usuario]
        propias.sort(key=lambda n: n.fecha_hora, reverse=True)
        return [n.to_dict() for n in propias]

    def contar_notificaciones_no_leidas(self, id_usuario):
        """Alimenta el badge de campanita en 'Subastas Activas' (ver
        views/subastas_activas_view.py), igual que
        contar_mensajes_no_leidos_totales alimenta el de mensajes."""
        return len([n for n in self.notificaciones
                    if n.id_usuario_destino == id_usuario and not n.leido])

    def marcar_notificaciones_leidas(self, id_usuario):
        """Se llama al abrir el panel de notificaciones -- mismo criterio que
        abrir un chat marca sus mensajes como leídos."""
        for n in self.notificaciones:
            if n.id_usuario_destino == id_usuario:
                n.leido = True

    def obtener_notificaciones_nuevas_sin_avisar(self, id_usuario):
        """Las que todavía no se le mostraron como aviso emergente (SnackBar)
        a ESTA cuenta dentro de la app. No exige que además estén sin leer —
        una notificación puede seguir sin leer después de avisada (ver
        docstring de Notificacion en modelos.py)."""
        nuevas = [n for n in self.notificaciones
                  if n.id_usuario_destino == id_usuario and not n.avisada_en_app]
        nuevas.sort(key=lambda n: n.fecha_hora)
        return [n.to_dict() for n in nuevas]

    def marcar_notificaciones_avisadas(self, id_usuario):
        """Se llama apenas se le muestra el SnackBar a la cuenta activa (ver
        main.py: avisar_notificaciones_nuevas), para no repetir el mismo
        aviso en la próxima reconstrucción de pantalla."""
        for n in self.notificaciones:
            if n.id_usuario_destino == id_usuario:
                n.avisada_en_app = True

    def obtener_todas_notificaciones_sin_avisar(self):
        """Versión GLOBAL (todos los usuarios) de
        obtener_notificaciones_nuevas_sin_avisar, para que main.py decida a
        quién avisarle por correo sin tener que consultar usuario por usuario
        (ver main.py: procesar_notificaciones_pendientes /
        backend/notificador_email.py)."""
        pendientes = [n for n in self.notificaciones if not n.avisada_en_app]
        pendientes.sort(key=lambda n: n.fecha_hora)
        return [n.to_dict() for n in pendientes]

    def marcar_notificacion_avisada(self, id_notificacion):
        """Versión por-id (una sola notificación), para el camino de correo:
        procesar_notificaciones_pendientes recorre notificaciones de
        distintos usuarios a la vez, así que necesita marcar de a una."""
        for n in self.notificaciones:
            if n.id == id_notificacion:
                n.avisada_en_app = True
                return

    # =====================================================================
    # MÓDULO VENTA (cierre de subastas)
    # =====================================================================
    def cerrar_venta_subasta(self, id_carro):
        carro = self.carros.get(id_carro)
        if not carro:
            return False, "El carro no existe."
        if carro.estado_subasta != "activa":
            return False, f"El carro ya no está activo (estado actual: {carro.estado_subasta})."

        pujas_carro = [p for p in self.pujas if p.id_carro == id_carro and p.pago_verificado]

        if not pujas_carro:
            carro.estado_subasta = "no_vendido"
            return True, None

        pujas_carro.sort(key=lambda x: x.monto, reverse=True)
        mejor_puja = pujas_carro[0]

        if mejor_puja.monto < carro.precio_reserva:
            carro.estado_subasta = "no_vendido"
            return True, None

        carro.estado_subasta = "vendido"
        carro.precio_final_venta = mejor_puja.monto
        carro.comprador_id = mejor_puja.id_usuario

        vendedor = self.usuarios.get(carro.vendedor_id)
        comprador = self.usuarios.get(mejor_puja.id_usuario)
        if vendedor and id_carro in vendedor.autos_en_posesion:
            vendedor.autos_en_posesion.remove(id_carro)
        if comprador:
            comprador.autos_en_posesion.append(id_carro)

        # Actualiza el historial personal de cada postor: Ganada / Perdida
        for p in pujas_carro:
            postor = self.usuarios.get(p.id_usuario)
            if not postor:
                continue
            for oferta in postor.historial_ofertas:
                if oferta["id_carro"] == id_carro:
                    oferta["estado"] = "Ganada" if p.id_usuario == mejor_puja.id_usuario else "Perdida"

        return True, carro

    def cerrar_subastas_vencidas(self):
        """Recorre todos los carros activos y cierra los que ya pasaron su fecha_fin.
        Esto antes no existía: el cierre era 100% manual."""
        cerrados = []
        for carro in list(self.carros.values()):
            if carro.estado_subasta == "activa" and carro.esta_vencida():
                self.cerrar_venta_subasta(carro.id)
                cerrados.append(carro.id)
        return cerrados

    # =====================================================================
    # CONFIRMACIÓN DE ENTREGA (post-venta)
    # =====================================================================
    def confirmar_entrega(self, id_carro, id_usuario):
        """
        El COMPRADOR ganador confirma que ya recibió el vehículo en la vida
        real. Solo tiene sentido para subastas ya cerradas como 'vendido', y
        solo el comprador puede confirmarlo (no el vendedor): es una
        constancia de que la contraparte que recibió el bien está conforme,
        no una autodeclaración del vendedor de que "ya lo entregó".

        No hay una acción simétrica para "deshacer" la confirmación a
        propósito — una vez confirmada la entrega, se considera un hecho
        consumado (igual de irreversible que el resto de los cierres de
        subasta en este sistema).
        """
        carro = self.carros.get(id_carro)
        if not carro:
            return False, "El carro no existe."
        if carro.estado_subasta != "vendido":
            return False, "Esta subasta todavía no está cerrada como vendida."
        if carro.comprador_id != id_usuario:
            return False, "Solo el comprador que ganó la subasta puede confirmar la entrega."
        if carro.entrega_confirmada:
            return False, "La entrega de este vehículo ya había sido confirmada."

        carro.entrega_confirmada = True
        carro.fecha_entrega_confirmada = datetime.now(timezone.utc).isoformat()
        return True, carro

    # =====================================================================
    # MÓDULO CALIFICACIONES (reseña del comprador sobre el vendedor)
    # =====================================================================
    def calificar_vendedor(self, id_carro, id_calificador, estrellas, comentario=""):
        """
        El COMPRADOR que ganó la subasta califica al vendedor. Solo se puede
        una vez que: la venta está cerrada ('vendido'), el propio comprador
        ya confirmó la entrega (ver confirmar_entrega arriba -- tiene que
        haber recibido el auto antes de opinar sobre la experiencia), y no
        haya calificado esa misma compra antes. No hay edición ni borrado de
        una calificación ya hecha, mismo criterio de "hecho consumado" que el
        resto de los cierres de este sistema.

        estrellas: entero de 1 a 5. La reputación del vendedor (ver
        Usuario.reputacion) se recalcula como el promedio de TODAS las
        calificaciones que recibió hasta ahora -- no un promedio
        incremental, para no arrastrar errores de redondeo.
        """
        carro = self.carros.get(id_carro)
        if not carro:
            return False, "El carro no existe."
        if carro.estado_subasta != "vendido":
            return False, "Esta subasta todavía no está cerrada como vendida."
        if carro.comprador_id != id_calificador:
            return False, "Solo el comprador que ganó la subasta puede calificar al vendedor."
        if not carro.entrega_confirmada:
            return False, "Tenés que confirmar la entrega antes de calificar al vendedor."
        try:
            estrellas = int(estrellas)
        except (TypeError, ValueError):
            return False, "La calificación tiene que ser un número de 1 a 5."
        if not (1 <= estrellas <= 5):
            return False, "La calificación tiene que ser de 1 a 5 estrellas."
        if self.ya_califico_compra(id_carro, id_calificador):
            return False, "Ya calificaste esta compra."

        vendedor = self.usuarios.get(carro.vendedor_id)
        if not vendedor:
            return False, "El vendedor ya no existe."

        nueva = Calificacion(
            id_calificacion=f"cal_{len(self.calificaciones) + 1:05d}",
            id_carro=id_carro,
            id_calificador=id_calificador,
            id_calificado=carro.vendedor_id,
            estrellas=estrellas,
            comentario=(comentario or "").strip(),
        )
        self.calificaciones.append(nueva)

        propias = [c.estrellas for c in self.calificaciones if c.id_calificado == carro.vendedor_id]
        vendedor.reputacion = round(sum(propias) / len(propias), 1)

        return True, nueva

    def ya_califico_compra(self, id_carro, id_usuario):
        """Para que la vista sepa si mostrar el formulario de calificar o un
        aviso de 'ya calificaste esta compra' (ver detalle_subasta_dialog.py)."""
        return any(c.id_carro == id_carro and c.id_calificador == id_usuario for c in self.calificaciones)

    def obtener_calificaciones_usuario(self, id_usuario):
        """Todas las calificaciones que recibió este usuario como vendedor,
        más recientes primero -- para mostrarlas en su perfil público (ver
        views/perfil_vendedor_dialog.py)."""
        propias = [c for c in self.calificaciones if c.id_calificado == id_usuario]
        propias.sort(key=lambda c: c.fecha_hora, reverse=True)
        resultado = []
        for c in propias:
            calificador = self.usuarios.get(c.id_calificador)
            d = c.to_dict()
            d["calificador_nombre"] = calificador.nombre if calificador else "Usuario"
            resultado.append(d)
        return resultado

    # =====================================================================
    # MÓDULO FAVORITOS
    # =====================================================================
    def agregar_favorito(self, id_usuario, id_carro):
        """
        Wrapper de sistema.py sobre Usuario.agregar_favorito() (que ya
        existía en modelos.py pero nunca se llamaba desde ninguna vista):
        agrega validación de existencia y el contrato (ok, resultado)
        estándar del resto de los métodos de acción de esta clase.
        """
        usuario = self.usuarios.get(id_usuario)
        if not usuario:
            return False, "El usuario no existe."
        if id_carro not in self.carros:
            return False, "El carro no existe."
        agregado = usuario.agregar_favorito(id_carro)
        if not agregado:
            return False, "Ese carro ya estaba en tus favoritos."
        return True, usuario

    def quitar_favorito(self, id_usuario, id_carro):
        usuario = self.usuarios.get(id_usuario)
        if not usuario:
            return False, "El usuario no existe."
        quitado = usuario.quitar_favorito(id_carro)
        if not quitado:
            return False, "Ese carro no estaba en tus favoritos."
        return True, usuario

    # =====================================================================
    # PUBLICACIÓN simplificada (lo que usa la pantalla "Mis Carros")
    # =====================================================================
    def publicar_carro(self, id_vendedor, marca, modelo, anio, kilometraje,
                        precio_base, precio_reserva, dias_duracion=7,
                        especificaciones=None, extras=None, imagen=None, imagenes=None,
                        condicion_general=None, descripcion_danos="",
                        documentos_en_regla=False):
        """
        Wrapper sobre recibir_carro_compra() que genera el id automáticamente,
        para que la pantalla de publicación solo tenga que pedir los campos
        que el vendedor realmente puede llenar. fecha_inicio/fecha_fin ya NO
        se calculan aquí: se asignan cuando un admin aprueba la subasta
        (ver aprobar_subasta), por eso solo guardamos dias_duracion por ahora.
        """
        coche_id = f"auto_{len(self.carros) + 1:03d}"
        return self.recibir_carro_compra(
            id_vendedor=id_vendedor, coche_id=coche_id, marca=marca, modelo=modelo,
            anio=anio, kilometraje=kilometraje, precio_base=precio_base,
            precio_reserva=precio_reserva, especificaciones=especificaciones or {},
            extras=extras or [], imagen=imagen, imagenes=imagenes, condicion_general=condicion_general,
            descripcion_danos=descripcion_danos, documentos_en_regla=documentos_en_regla,
            duracion_dias=dias_duracion,
        )

    # =====================================================================
    # CONSULTAS PARA LAS PESTAÑAS DEL FRONT END
    # =====================================================================
    def obtener_puja_maxima(self, id_carro):
        """Monto más alto pujado por un carro (o su precio_base si no hay pujas)."""
        carro = self.carros.get(id_carro)
        if not carro:
            return 0
        pujas_validas = [p.monto for p in self.pujas if p.id_carro == id_carro and p.pago_verificado]
        return max([carro.precio_base] + pujas_validas)

    def obtener_carro_por_id(self, id_carro):
        """
        Dict de un solo carro por id (via Carro.to_dict()), o None si no
        existe. La mayoría de las vistas ya reciben el dict del carro armado
        por otro "obtener_..." (obtener_mis_carros, obtener_subastas_explorar,
        etc.), pero la bandeja de mensajes global (ver
        obtener_todas_mis_conversaciones más abajo) solo conoce el id_carro
        de cada conversación y necesita reconstruir el dict completo para
        poder abrir el chat correspondiente sin tocar el objeto Carro directo.
        """
        carro = self.carros.get(id_carro)
        return carro.to_dict() if carro else None

    def obtener_mis_carros(self, id_usuario):
        """Pestaña 'MIS CARROS': todo lo que este usuario ha publicado, sea cual sea su estado."""
        mios = [c for c in self.carros.values() if c.vendedor_id == id_usuario]
        mios.sort(key=lambda c: (c.estado_subasta != "activa", c.fecha_fin or ""))
        resultado = []
        for c in mios:
            d = c.to_dict()
            d["puja_maxima"] = self.obtener_puja_maxima(c.id)
            d["num_pujas"] = len([p for p in self.pujas if p.id_carro == c.id])
            resultado.append(d)
        return resultado

    def obtener_subastas_explorar(self, id_usuario=None, filtro_texto=None, marca=None,
                                   precio_min=None, precio_max=None, anio_min=None, anio_max=None,
                                   orden="cierra_pronto"):
        """
        Pestaña 'EXPLORAR SUBASTAS': todas las subastas activas de la plataforma.

        filtro_texto: si se pasa, solo se devuelven los carros cuya marca,
        modelo o año contienen ese texto (sin distinguir mayúsculas/acentos
        exactos — comparación simple, suficiente para el buscador de la
        barra superior).

        Filtros adicionales (todos opcionales, se combinan con AND entre sí
        y con filtro_texto):
          marca: coincidencia exacta (ya viene de un dropdown poblado con
                 obtener_marcas_activas(), no hace falta comparación difusa).
          precio_min / precio_max: sobre la PUJA MÁS ALTA actual (lo que de
                 verdad costaría ganar la subasta ahora mismo), no sobre
                 precio_base.
          anio_min / anio_max: sobre el año del vehículo.
          orden: 'cierra_pronto' (default, como antes) | 'precio_asc' |
                 'precio_desc' | 'mas_pujas'.
        """
        activos = [c for c in self.carros.values() if c.estado_subasta == "activa"]
        if filtro_texto:
            q = filtro_texto.strip().lower()
            activos = [
                c for c in activos
                if q in c.marca.lower() or q in c.modelo.lower() or q in str(c.anio)
            ]
        if marca:
            activos = [c for c in activos if c.marca == marca]
        if anio_min is not None:
            activos = [c for c in activos if c.anio >= anio_min]
        if anio_max is not None:
            activos = [c for c in activos if c.anio <= anio_max]

        resultado = []
        for c in activos:
            d = c.to_dict()
            d["puja_maxima"] = self.obtener_puja_maxima(c.id)
            d["num_pujas"] = len([p for p in self.pujas if p.id_carro == c.id])
            d["es_propio"] = (c.vendedor_id == id_usuario) if id_usuario else False
            resultado.append(d)

        if precio_min is not None:
            resultado = [d for d in resultado if d["puja_maxima"] >= precio_min]
        if precio_max is not None:
            resultado = [d for d in resultado if d["puja_maxima"] <= precio_max]

        claves_orden = {
            "cierra_pronto": lambda d: d["horas_restantes"] if d["horas_restantes"] is not None else float("inf"),
            "precio_asc": lambda d: d["puja_maxima"],
            "precio_desc": lambda d: -d["puja_maxima"],
            "mas_pujas": lambda d: -d["num_pujas"],
        }
        resultado.sort(key=claves_orden.get(orden, claves_orden["cierra_pronto"]))
        return resultado

    def obtener_marcas_activas(self):
        """Marcas distintas entre las subastas activas, para poblar el
        dropdown de filtro de 'Explorar Subastas' sin inventar una lista fija
        de marcas — si mañana se publica una marca nueva, aparece sola."""
        return sorted({c.marca for c in self.carros.values() if c.estado_subasta == "activa"})

    def obtener_mis_subastas_activas(self, id_usuario):
        """
        Pestaña 'SUBASTAS ACTIVAS': subastas activas en las que este usuario
        participa (tiene una puja o la marcó como favorita). Es justo la
        diferencia con 'Explorar Subastas', que muestra TODO el mercado.
        """
        usuario = self.usuarios.get(id_usuario)
        if not usuario:
            return []  # usuario inexistente: no hay nada que mostrarle

        ids_con_oferta = {o["id_carro"] for o in usuario.historial_ofertas}
        ids_favoritos = set(usuario.favoritos)
        ids_interes = ids_con_oferta | ids_favoritos

        activos = [c for c in self.carros.values()
                   if c.estado_subasta == "activa" and c.id in ids_interes]
        activos.sort(key=lambda c: c.tiempo_restante() or 0)

        resultado = []
        for c in activos:
            mi_oferta = next((o for o in usuario.historial_ofertas if o["id_carro"] == c.id), None)
            d = c.to_dict()
            d["puja_maxima"] = self.obtener_puja_maxima(c.id)
            d["mi_estado"] = mi_oferta["estado"] if mi_oferta else "Solo en favoritos"
            d["mi_monto"] = mi_oferta["monto"] if mi_oferta else None
            resultado.append(d)
        return resultado

    def obtener_mis_ventas(self, id_usuario):
        """Pestaña 'VENTAS': historial de subastas YA CERRADAS de este vendedor."""
        cerradas = [c for c in self.carros.values()
                    if c.vendedor_id == id_usuario and c.estado_subasta in ("vendido", "no_vendido")]
        cerradas.sort(key=lambda c: c.fecha_fin or "", reverse=True)

        resultado = []
        for c in cerradas:
            comprador = self.usuarios.get(c.comprador_id) if c.comprador_id else None
            d = c.to_dict()
            d["comprador_nombre"] = comprador.nombre if comprador else None
            d["comision"] = c.precio_final_venta * self.comision_porcentaje if c.estado_subasta == "vendido" else 0
            resultado.append(d)
        return resultado

    def obtener_mis_compras_ganadas(self, id_usuario):
        """
        Carros que este usuario GANÓ (es el comprador de una subasta ya
        vendida), sin importar si la entrega ya se confirmó o no. Es la
        fuente de datos "cruda" para la sección de 'Mis Gastos' en la
        pestaña 'VENTAS' (que quiere ver el historial completo de compras,
        entregadas o no, porque ahí el foco es el dinero gastado).

        Para las otras dos pantallas que separan por estado de entrega, usar
        los métodos de abajo en vez de filtrar a mano sobre este resultado:
          - obtener_mis_compras_pendientes_entrega() -> 'SUBASTAS ACTIVAS'
          - obtener_mis_carros_ganados()              -> 'MIS CARROS'
        """
        ganadas = [c for c in self.carros.values()
                   if c.comprador_id == id_usuario and c.estado_subasta == "vendido"]
        ganadas.sort(key=lambda c: c.fecha_fin or "", reverse=True)

        resultado = []
        for c in ganadas:
            vendedor = self.usuarios.get(c.vendedor_id)
            d = c.to_dict()
            d["vendedor_nombre"] = vendedor.nombre if vendedor else "Desconocido"
            resultado.append(d)
        return resultado

    def obtener_mis_compras_pendientes_entrega(self, id_usuario):
        """
        Subconjunto de obtener_mis_compras_ganadas(): solo las que este
        usuario ganó como comprador y TODAVÍA no confirmó haber recibido en
        la vida real (ver confirmar_entrega). Es lo que alimenta 'SUBASTAS
        ACTIVAS' → 'Compras ganadas' — ahí es donde tiene sentido seguir
        coordinando la entrega por chat. En cuanto el comprador confirma que
        ya la recibió, el carro deja de aparecer acá y pasa a listarse en
        'MIS CARROS' → 'Carros Ganados' (ver obtener_mis_carros_ganados).
        """
        return [c for c in self.obtener_mis_compras_ganadas(id_usuario) if not c.get("entrega_confirmada")]

    def obtener_mis_carros_ganados(self, id_usuario):
        """
        Subconjunto de obtener_mis_compras_ganadas(): solo las que este
        usuario ganó como comprador Y cuya entrega YA confirmó (ver
        confirmar_entrega). Alimenta 'MIS CARROS' → 'Carros Ganados': hasta
        que no se confirma la entrega, el vehículo se sigue viendo
        únicamente en 'SUBASTAS ACTIVAS' → 'Compras ganadas' (ver
        obtener_mis_compras_pendientes_entrega arriba); recién cuando el
        comprador marca la casilla/botón de entrega, el carro pasa a
        contarse también como "suyo" acá, en Mis Carros — junto con los
        que este usuario publicó como vendedor.
        """
        return [c for c in self.obtener_mis_compras_ganadas(id_usuario) if c.get("entrega_confirmada")]

    # =====================================================================
    # PERFIL PÚBLICO (para que un comprador evalúe si un vendedor es confiable)
    # =====================================================================
    def obtener_perfil_publico_usuario(self, id_usuario):
        """
        Info pública de un usuario, pensada para mostrarse a CUALQUIERA que
        haga clic en su nombre desde el detalle de una subasta (no solo a
        él mismo). Por eso NO incluye email, teléfono ni password_hash —
        eso es privado y ya se muestra aparte en perfil_view.py para el
        dueño de la cuenta. foto_perfil SÍ se incluye a propósito: es
        información pública por diseño (se ve en las tarjetas de subastas
        de cualquiera), igual que el nombre o la reputación.

        Devuelve None si el usuario no existe (ej. datos borrados/cambiados),
        igual que el resto de los "obtener_..." tolera referencias rotas sin
        levantar excepción — el llamador decide qué mostrar en ese caso.
        """
        usuario = self.usuarios.get(id_usuario)
        if not usuario:
            return None

        publicados = [c for c in self.carros.values() if c.vendedor_id == id_usuario]
        vendidos = [c for c in publicados if c.estado_subasta == "vendido"]
        activos = [c for c in publicados if c.estado_subasta == "activa"]

        return {
            "id": usuario.id,
            "nombre": usuario.nombre,
            "foto_perfil": usuario.foto_perfil,
            "verificado": usuario.verificado,
            "reputacion": usuario.reputacion,
            "fecha_registro": usuario.fecha_registro,
            "autos_publicados": len(publicados),
            "autos_vendidos": len(vendidos),
            "subastas_activas": len(activos),
        }

    # =====================================================================
    # MÓDULO CHAT (contacto comprador-vendedor)
    # =====================================================================
    def enviar_mensaje(self, id_carro, id_remitente, id_destinatario, texto):
        if id_carro not in self.carros:
            return False, "El carro no existe."
        if id_remitente not in self.usuarios or id_destinatario not in self.usuarios:
            return False, "Uno de los usuarios no existe."
        if id_remitente == id_destinatario:
            return False, "No puedes enviarte un mensaje a ti mismo."
        texto = (texto or "").strip()
        if not texto:
            return False, "El mensaje no puede estar vacío."

        nuevo = Mensaje(
            id_mensaje=f"msg_{len(self.mensajes) + 1:05d}",
            id_carro=id_carro,
            id_remitente=id_remitente,
            id_destinatario=id_destinatario,
            texto=texto,
        )
        self.mensajes.append(nuevo)
        return True, nuevo

    def obtener_conversacion(self, id_carro, id_usuario_a, id_usuario_b):
        """Todos los mensajes entre estas dos personas sobre este carro, en orden cronológico."""
        participantes = {id_usuario_a, id_usuario_b}
        conversacion = [
            m for m in self.mensajes
            if m.id_carro == id_carro and {m.id_remitente, m.id_destinatario} == participantes
        ]
        conversacion.sort(key=lambda m: m.fecha_hora)
        return conversacion

    def marcar_conversacion_leida(self, id_carro, id_usuario_a, id_usuario_b, lector_id):
        """Marca como leídos los mensajes de esta conversación dirigidos a lector_id."""
        for m in self.obtener_conversacion(id_carro, id_usuario_a, id_usuario_b):
            if m.id_destinatario == lector_id:
                m.leido = True

    def obtener_conversaciones_carro(self, id_carro, id_vendedor):
        """
        Para el vendedor: lista de personas distintas que le han escrito (o a
        quienes les escribió) sobre este carro, con el último mensaje y
        cuántos quedan sin leer. Útil porque varios postores pueden escribir
        sobre el mismo carro y no deben mezclarse en un solo hilo.

        Esta vista es POR CARRO y pensada para el detalle de una subasta
        puntual (views/detalle_subasta_dialog.py). Para ver TODAS las
        conversaciones del usuario en toda la plataforma de una sola vez
        (sin importar el carro, ni si escribió como vendedor o comprador),
        ver obtener_todas_mis_conversaciones() más abajo.
        """
        otros_ids = set()
        for m in self.mensajes:
            if m.id_carro != id_carro:
                continue
            if m.id_remitente == id_vendedor:
                otros_ids.add(m.id_destinatario)
            elif m.id_destinatario == id_vendedor:
                otros_ids.add(m.id_remitente)

        resultado = []
        for otro_id in otros_ids:
            otro = self.usuarios.get(otro_id)
            if not otro:
                continue
            conversacion = self.obtener_conversacion(id_carro, id_vendedor, otro_id)
            if not conversacion:
                continue
            ultimo = conversacion[-1]
            no_leidos = len([m for m in conversacion if m.id_destinatario == id_vendedor and not m.leido])
            resultado.append({
                "otro_usuario_id": otro_id,
                "otro_usuario_nombre": otro.nombre,
                "ultimo_mensaje": ultimo.texto,
                "fecha_ultimo_mensaje": ultimo.fecha_hora,
                "no_leidos": no_leidos,
            })
        resultado.sort(key=lambda r: r["fecha_ultimo_mensaje"], reverse=True)
        return resultado

    def contar_mensajes_no_leidos_carro(self, id_carro, id_usuario):
        """Total de mensajes sin leer dirigidos a id_usuario sobre este carro (cualquier remitente)."""
        return len([m for m in self.mensajes
                    if m.id_carro == id_carro and m.id_destinatario == id_usuario and not m.leido])

    def obtener_todas_mis_conversaciones(self, id_usuario):
        """
        TODAS las conversaciones de este usuario en toda la plataforma, sin
        importar el carro ni si escribió como vendedor o como comprador —a
        diferencia de obtener_conversaciones_carro(), que es por-carro y
        solo tiene sentido desde el punto de vista del vendedor de ESE carro
        puntual. Esta es la fuente de datos de la bandeja de mensajes global
        (ver views/bandeja_mensajes_dialog.py), a la que se accede desde el
        ícono de mensajes en la barra superior (views/shared.py: top_bar).

        Cada conversación queda identificada por (id_carro, otro_usuario_id),
        igual que en obtener_conversacion(). Si el carro o el otro usuario ya
        no existen (datos borrados/cambiados), esa conversación simplemente
        se omite del resultado — mismo criterio tolerante que el resto de
        los métodos "obtener_...".

        Se ordena por fecha del último mensaje, más reciente primero (igual
        que cualquier bandeja de chat convencional).
        """
        claves_vistas = set()
        claves_conversaciones = []
        for m in self.mensajes:
            if m.id_remitente == id_usuario:
                otro_id = m.id_destinatario
            elif m.id_destinatario == id_usuario:
                otro_id = m.id_remitente
            else:
                continue
            clave = (m.id_carro, otro_id)
            if clave not in claves_vistas:
                claves_vistas.add(clave)
                claves_conversaciones.append(clave)

        resultado = []
        for id_carro, otro_id in claves_conversaciones:
            otro = self.usuarios.get(otro_id)
            carro = self.carros.get(id_carro)
            if not otro or not carro:
                continue

            conversacion = self.obtener_conversacion(id_carro, id_usuario, otro_id)
            if not conversacion:
                continue
            ultimo = conversacion[-1]
            no_leidos = len([m for m in conversacion if m.id_destinatario == id_usuario and not m.leido])

            resultado.append({
                "id_carro": id_carro,
                "carro_marca": carro.marca,
                "carro_modelo": carro.modelo,
                "carro_anio": carro.anio,
                "carro_imagen": carro.imagen,
                "otro_usuario_id": otro_id,
                "otro_usuario_nombre": otro.nombre,
                "otro_usuario_foto": otro.foto_perfil,
                "ultimo_mensaje": ultimo.texto,
                "ultimo_mensaje_es_mio": ultimo.id_remitente == id_usuario,
                "fecha_ultimo_mensaje": ultimo.fecha_hora,
                "no_leidos": no_leidos,
                "soy_vendedor_del_carro": carro.vendedor_id == id_usuario,
            })

        resultado.sort(key=lambda r: r["fecha_ultimo_mensaje"], reverse=True)
        return resultado

    def contar_mensajes_no_leidos_totales(self, id_usuario):
        """
        Total de mensajes sin leer dirigidos a id_usuario en TODA la
        plataforma (todas las conversaciones, todos los carros). Es la
        versión "global" de contar_mensajes_no_leidos_carro() (que solo
        cuenta los de un carro puntual) y alimenta el badge del ícono de
        mensajes en la barra superior (ver views/shared.py: top_bar /
        page_shell, que la calcula en cada re-render de la pantalla).
        """
        return len([m for m in self.mensajes if m.id_destinatario == id_usuario and not m.leido])

    # =====================================================================
    # MÉTRICAS / REPORTES (pensados para alimentar el dashboard de Flet)
    # =====================================================================
    def calcular_resumen_financiero(self):
        vendidos = [c for c in self.carros.values() if c.estado_subasta == "vendido"]
        total_ventas = sum(c.precio_final_venta for c in vendidos)
        ganancia_plataforma = total_ventas * self.comision_porcentaje
        return {
            "autos_catalogo": len(self.carros),
            "autos_vendidos": len(vendidos),
            "volumen_total": total_ventas,
            "comision_porcentaje": self.comision_porcentaje,
            "ganancia_plataforma": ganancia_plataforma,
        }

    def generar_reporte_financiero(self):
        """Versión CLI del reporte (se mantiene por compatibilidad con admin.py original)."""
        r = self.calcular_resumen_financiero()
        print("\n==================================================")
        print("    REPORTE FINANCIERO DE COMPRA Y VENTA (ADMIN)   ")
        print("==================================================")
        print(f" Autos en catálogo total:     {r['autos_catalogo']}")
        print(f" Autos vendidos con éxito:   {r['autos_vendidos']}")
        print(f" Volumen Total Transaccionado: ${r['volumen_total']:,.2f}")
        print(f" Comisión de Plataforma ({r['comision_porcentaje']*100}%): ${r['ganancia_plataforma']:,.2f}")
        print("==================================================\n")

    def obtener_resumen_dashboard(self, id_usuario):
        """
        Números para las 3 tarjetas superiores del dashboard (RESUMEN), vistas
        desde la perspectiva de un usuario específico.

        Definiciones (decisión de producto, ajustar si el equipo define otra cosa):
        - ganancias: suma de precio_final_venta de los autos que este usuario VENDIÓ.
        - gastado: suma de precio_final_venta de los autos que este usuario COMPRÓ.
        - subastas_activas_pendientes: subastas activas en TODA la plataforma
          (no solo las del usuario), igual que se ve en la referencia de diseño.
        """
        vendidos_por_el = [c for c in self.carros.values()
                            if c.vendedor_id == id_usuario and c.estado_subasta == "vendido"]
        comprados_por_el = [c for c in self.carros.values()
                             if c.comprador_id == id_usuario and c.estado_subasta == "vendido"]
        activas = [c for c in self.carros.values() if c.estado_subasta == "activa"]

        return {
            "ganancias": sum(c.precio_final_venta for c in vendidos_por_el),
            "gastado": sum(c.precio_final_venta for c in comprados_por_el),
            "subastas_activas_pendientes": len(activas),
        }

    def obtener_subastadores_frecuentes(self, top_n=5):
        """Usuarios que más pujas han hecho en toda la plataforma."""
        conteo = {}
        for p in self.pujas:
            conteo[p.id_usuario] = conteo.get(p.id_usuario, 0) + 1

        ranking = sorted(conteo.items(), key=lambda kv: kv[1], reverse=True)[:top_n]
        resultado = []
        for id_usuario, cantidad in ranking:
            usuario = self.usuarios.get(id_usuario)
            if usuario:
                resultado.append({
                    "nombre": usuario.nombre,
                    "email": usuario.email,
                    "cantidad_pujas": cantidad,
                })
        return resultado

    def obtener_subastas_por_cerrar(self, top_n=5):
        """
        Reemplaza a la tarjeta 'TASAS DE CAMBIO' del mockup (esa era contenido
        de plantilla de Figma, no tiene relación con el negocio). Esta sí es
        información real y útil: las subastas activas más próximas a cerrar.
        """
        activas = [c for c in self.carros.values()
                   if c.estado_subasta == "activa" and c.tiempo_restante() is not None]
        activas.sort(key=lambda c: c.tiempo_restante())

        resultado = []
        for carro in activas[:top_n]:
            num_pujas = len([p for p in self.pujas if p.id_carro == carro.id])
            horas = carro.tiempo_restante().total_seconds() / 3600
            resultado.append({
                "carro": f"{carro.marca} {carro.modelo}",
                "num_pujas": num_pujas,
                "horas_restantes": round(horas, 1),
            })
        return resultado

    def obtener_ingresos_mensuales(self):
        """
        Agrega precio_final_venta por mes a partir de las ventas reales.

        NOTA: con la cantidad de datos de ejemplo que hay ahora (1-2 ventas),
        esto va a devolver muy pocos puntos. Es el comportamiento correcto:
        el gráfico de ingresos del dashboard se va a ver disperso hasta que
        haya más historial real de ventas. No hay que rellenarlo con datos
        inventados.
        """
        por_mes = {}
        for c in self.carros.values():
            if c.estado_subasta == "vendido" and c.fecha_fin:
                mes = c.fecha_fin[:7]  # 'YYYY-MM'
                por_mes[mes] = por_mes.get(mes, 0) + c.precio_final_venta
        return dict(sorted(por_mes.items()))

    def obtener_actividad_pujas_mensual(self):
        """
        Cantidad de pujas por mes (toda la plataforma). Es la métrica que
        alimenta 'MOVIMIENTOS NETOS' en el dashboard: una cosa distinta al
        volumen de ventas, para no graficar dos veces el mismo número.
        Mismo aviso que en obtener_ingresos_mensuales(): con pocos datos de
        ejemplo, va a haber pocos puntos. Es esperado.
        """
        por_mes = {}
        for p in self.pujas:
            if p.fecha_hora:
                mes = p.fecha_hora[:7]
                por_mes[mes] = por_mes.get(mes, 0) + 1
        return dict(sorted(por_mes.items()))

    # =====================================================================
    # REPORTE FINANCIERO PERSONAL (ingresos como vendedor / gastos como
    # comprador), para la pestaña 'VENTAS' y su exportación a Excel.
    # =====================================================================
    def obtener_movimientos_financieros(self, id_usuario):
        """
        Historial combinado de TODOS los movimientos de dinero de este
        usuario en la plataforma: cada venta cerrada donde fue vendedor
        cuenta como un 'ingreso' (con la comisión de la plataforma ya
        descontada en 'monto_neto'), y cada compra ganada donde fue
        comprador cuenta como un 'egreso'. Solo subastas 'vendido' generan
        movimiento — 'no_vendido'/'rechazada' no mueven dinero de nadie.

        Pensado como la fuente única de verdad tanto para la sección de
        gastos en la pestaña VENTAS como para el reporte Excel: ambos
        consumen esta misma lista para no duplicar la lógica de cálculo.
        """
        movimientos = []
        for c in self.carros.values():
            if c.estado_subasta != "vendido":
                continue
            if c.vendedor_id == id_usuario:
                comision = c.precio_final_venta * self.comision_porcentaje
                movimientos.append({
                    "fecha": c.fecha_fin,
                    "tipo": "ingreso",
                    "concepto": f"Venta: {c.marca} {c.modelo} ({c.anio})",
                    "monto": c.precio_final_venta,
                    "comision": comision,
                    "monto_neto": c.precio_final_venta - comision,
                    "carro_id": c.id,
                    "contraparte_id": c.comprador_id,
                })
            if c.comprador_id == id_usuario:
                movimientos.append({
                    "fecha": c.fecha_fin,
                    "tipo": "egreso",
                    "concepto": f"Compra: {c.marca} {c.modelo} ({c.anio})",
                    "monto": c.precio_final_venta,
                    "comision": 0.0,
                    "monto_neto": c.precio_final_venta,
                    "carro_id": c.id,
                    "contraparte_id": c.vendedor_id,
                })
        movimientos.sort(key=lambda m: m["fecha"] or "")
        return movimientos

    def obtener_resumen_financiero_usuario(self, id_usuario):
        """Totales agregados a partir de obtener_movimientos_financieros(), para
        las tarjetas de resumen de la pestaña VENTAS y la hoja 'Resumen' del Excel."""
        movimientos = self.obtener_movimientos_financieros(id_usuario)
        ingresos = [m for m in movimientos if m["tipo"] == "ingreso"]
        egresos = [m for m in movimientos if m["tipo"] == "egreso"]
        total_ingresos = sum(m["monto"] for m in ingresos)
        total_comisiones = sum(m["comision"] for m in ingresos)
        total_egresos = sum(m["monto"] for m in egresos)
        return {
            "total_ingresos": total_ingresos,
            "total_comisiones": total_comisiones,
            "ingresos_netos": total_ingresos - total_comisiones,
            "total_egresos": total_egresos,
            "balance_neto": (total_ingresos - total_comisiones) - total_egresos,
            "num_ventas": len(ingresos),
            "num_compras": len(egresos),
        }

    def obtener_movimientos_mensuales_usuario(self, id_usuario):
        """Ingresos netos y egresos agrupados por mes, para graficar (tanto en
        el gráfico de barras del Excel como, a futuro, en algún gráfico inline)."""
        por_mes = {}
        for m in self.obtener_movimientos_financieros(id_usuario):
            mes = (m["fecha"] or "")[:7]
            if not mes:
                continue
            entrada = por_mes.setdefault(mes, {"ingresos": 0.0, "egresos": 0.0})
            if m["tipo"] == "ingreso":
                entrada["ingresos"] += m["monto_neto"]
            else:
                entrada["egresos"] += m["monto"]
        return dict(sorted(por_mes.items()))

    def generar_reporte_excel_usuario(self, id_usuario, ruta_destino):
        """
        Genera un archivo .xlsx con 3 hojas: 'Resumen' (totales), 'Movimientos'
        (detalle fila por fila de cada venta/compra) e 'Ingresos vs Egresos'
        (tabla mensual + gráfico de barras). Usa openpyxl (ver requirements.txt).

        Si openpyxl no está instalado, devuelve un error legible en vez de
        romper la app — la vista debe mostrar ese mensaje al usuario en vez
        de asumir que la función siempre puede generar el archivo.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            return False, "Falta instalar 'openpyxl' (pip install openpyxl) para generar el reporte Excel."

        usuario = self.usuarios.get(id_usuario)
        if not usuario:
            return False, "El usuario no existe."

        movimientos = self.obtener_movimientos_financieros(id_usuario)
        resumen = self.obtener_resumen_financiero_usuario(id_usuario)
        por_mes = self.obtener_movimientos_mensuales_usuario(id_usuario)

        wb = Workbook()
        titulo_font = Font(bold=True, size=14)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="13132B")

        # --- Hoja 1: Resumen ---
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen["A1"] = f"Reporte financiero — {usuario.nombre}"
        ws_resumen["A1"].font = titulo_font
        ws_resumen["A2"] = f"Generado el {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"

        filas_resumen = [
            ("Total ingresos (ventas)", resumen["total_ingresos"]),
            ("Comisiones de plataforma pagadas", resumen["total_comisiones"]),
            ("Ingresos netos", resumen["ingresos_netos"]),
            ("Total egresos (compras)", resumen["total_egresos"]),
            ("Balance neto", resumen["balance_neto"]),
            ("Número de ventas", resumen["num_ventas"]),
            ("Número de compras", resumen["num_compras"]),
        ]
        fila_inicio = 4
        for i, (etiqueta, valor) in enumerate(filas_resumen):
            ws_resumen.cell(row=fila_inicio + i, column=1, value=etiqueta)
            ws_resumen.cell(row=fila_inicio + i, column=2, value=valor)
        ws_resumen.column_dimensions["A"].width = 34
        ws_resumen.column_dimensions["B"].width = 18

        # --- Hoja 2: Movimientos detallados ---
        ws_mov = wb.create_sheet("Movimientos")
        encabezados = ["Fecha", "Tipo", "Concepto", "Monto", "Comisión", "Monto neto"]
        for col, encabezado in enumerate(encabezados, start=1):
            celda = ws_mov.cell(row=1, column=col, value=encabezado)
            celda.font = header_font
            celda.fill = header_fill

        for row, m in enumerate(movimientos, start=2):
            ws_mov.cell(row=row, column=1, value=(m["fecha"] or "")[:10])
            ws_mov.cell(row=row, column=2, value="Ingreso" if m["tipo"] == "ingreso" else "Egreso")
            ws_mov.cell(row=row, column=3, value=m["concepto"])
            ws_mov.cell(row=row, column=4, value=m["monto"])
            ws_mov.cell(row=row, column=5, value=m["comision"])
            ws_mov.cell(row=row, column=6, value=m["monto_neto"])

        for col, ancho in zip("ABCDEF", (12, 10, 36, 14, 12, 14)):
            ws_mov.column_dimensions[col].width = ancho

        # --- Hoja 3: Datos mensuales + gráfico de barras ---
        ws_chart = wb.create_sheet("Ingresos vs Egresos")
        ws_chart.append(["Mes", "Ingresos", "Egresos"])
        for celda in ws_chart[1]:
            celda.font = header_font
            celda.fill = header_fill

        for mes, datos in por_mes.items():
            ws_chart.append([mes, datos["ingresos"], datos["egresos"]])

        if por_mes:
            chart = BarChart()
            chart.title = "Ingresos vs Egresos por mes"
            chart.y_axis.title = "Monto"
            chart.x_axis.title = "Mes"
            num_filas = len(por_mes)
            datos_ref = Reference(ws_chart, min_col=2, max_col=3, min_row=1, max_row=1 + num_filas)
            categorias_ref = Reference(ws_chart, min_col=1, min_row=2, max_row=1 + num_filas)
            chart.add_data(datos_ref, titles_from_data=True)
            chart.set_categories(categorias_ref)
            chart.width = 20
            chart.height = 10
            ws_chart.add_chart(chart, "E2")
        else:
            ws_chart["A2"] = "Todavía no hay movimientos con fecha para graficar."

        for col, ancho in zip("ABC", (12, 14, 14)):
            ws_chart.column_dimensions[col].width = ancho

        try:
            wb.save(ruta_destino)
        except Exception as e:
            return False, f"No se pudo guardar el archivo: {e}"

        return True, ruta_destino
